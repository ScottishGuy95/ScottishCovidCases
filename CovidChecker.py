#! python3
# CovidChecker.py - A program to check the new coronavirus cases in Scotland
# Data is taken from - https://www.gov.scot/publications/coronavirus-covid-19-trends-in-daily-data/

# TODO: Setup a -h or help option to get details on how to use the program
# TODO: Remove unnecessary prints from WHOLE program
# TODO: Program needs an internet connection (for now, change in future version if I can)
#  So it should start, check for a file, if no file exists, try to get a fresh file. If it fails, let exception handle.
#  If there is a file, but its not with today's date, try to get a fresh file, if fails, exception handles it.
# pip installs
# bs4, requests, openpyxl, send2trash

# Imports
from bs4 import BeautifulSoup as bs
import requests
from datetime import date
import urllib.request
from openpyxl import *
import os
import sys
import shutil
from send2trash import send2trash
import argparse


# Functions
def getFormattedDate(formatted=True):
    """
    Returns the date in a specified format, depending on parameters
    :param formatted: If True, returns the data formatted for a URL, otherwise its just DDMonthYYYY
    :return: (str): A String representing the formatted data
    """
    # Get date in DDMMYYYY format
    today = date.today()  # Create an object to get the date details
    dayNum = today.strftime("%d")  # Returns today's date in a int
    month = today.strftime("%B")  # Returns today's month in word format
    year = today.strftime("%Y")  # Returns year in 4 digits
    # Check if formatting is required or not, and return the value as a String
    if formatted:
        # Returns data, formatted to be used as part of a URL
        dateStr = dayNum + "%2B" + month + "%2B" + year
    else:
        # Returns today's date in format - DDMonthYear, e.g. 25October2020
        dateStr = dayNum + month + year
    return dateStr


def downloadData(url, file):
    """
    Downloads the latest excel file of Scottish Covid cases from the given URL
    :param url: (str): The URL of the Scottish Gov Covid Cases website
    :param file: (str; The name of the file to check for
    """
    # Check if a file with that name already exists, if not, download a fresh copy
    # Uses the os module to check the name of the file to download, against all the files in the ExcelFiles directory
    if file in os.listdir(os.getcwd() + '\\ExcelFiles\\'):
        print('A file with today\'s date already exists, using that.')
    else:
        try:
            # Attempts to use the urllib module to download the given file from the Scot Gov website
            urllib.request.urlretrieve(url, file)  # Downloads to the same directory as the python file
        # If there is an issue with the given URL, display an error, the given URL and end the program
        except urllib.error.URLError:
            print("Error! The given URL is incorrect, please check the URL")
            print("URL Given: " + url)
            print('Ending program as no data available to analyse')
            sys.exit()  # Ends the program as the URL failed, so no data available
        print("Local data matches most recently available data")
        # Move the newly downloaded file into the correct directory
        # TODO: Add some sort of error in case moving the file returns an error
        shutil.move(newestFile, 'ExcelFiles')


def getURLs():
    """
    Scans a web page for all URLs
    :return: (list); A list of all of the URLS from the Scottish Gov Covid page
    """
    # The website that hosts the files
    url = 'https://www.gov.scot/publications/coronavirus-covid-19-trends-in-daily-data/'
    # Create a soup object, that reads the HTML content and create a list to later store the results
    soup = bs(requests.get(url).text, 'html.parser')
    links = []

    # Loop through all of the content in the soup object, adding each hyperlink to the list and return the list
    # This allows the hyperlinks to be searched for the correct file later, by 'etFileNameFromLinks()
    for a in soup.find_all('a'):
        links.append(a['href'])
    return links


def formatFileName():
    """
    Gets the name of the covid data sheet and formats it to rename the exhcel sheet when its downloaded
    :return: (str): A String to name the downloaded excel sheet of covid data
    """
    links = getURLs()  # Get a list of URLs to search through for the Excel file
    fileName = ''  # The intended filename for the download object

    try:
        # Loop through all of the URLs in the list, for the specific file name and store in the fileName variable
        for link in links:
            if '/binaries/content/documents/govscot/publications/statistics/2020/04/coronavirus-covid-19-trends-in' \
               '-daily-data/documents/covid-19-data-by-nhs-board/' in link:
                if '?forceDownload' in link:
                    fileName = link

        stringPos = fileName.index('COVID-19%2B')  # Finds the start of the file name and stores its character position
        # Removes the first part of the URL to get a cleaner file name, using the above index as the cut off point
        fileName = fileName[stringPos:]
        # Removes all the characters from the String that are not needed (cleaning it up)
        fileName = fileName.replace('%2B', '')
        fileName = fileName.replace('?forceDownload=true', '')
        fileName = fileName.replace('dailydata-byNHSBoard', '')
        # The file name in a readable format
        return fileName
    except:
        # Print warning message to user
        print('Error! Found no file to download. Please check the URL has a valid file to download.\nIt is possible '
              'the file name has changed.')
        sys.exit()


def getHealthBoardList():
    """
    Reads the covid data and returns all of the health boards
    :return: (list): A list of each Health Board name as a Strings
    """
    healthBoardsList = []                           # Used to store all the health board names
    # Loops from the first available health board name (col 2) to the last health board name (col 16)
    for row in sheet.iter_rows(min_row=3, min_col=2, max_row=3, max_col=16):
        for cell in row:                            # Get the cell of the iteration
            healthBoardsList.append(cell.value)     # Add the health board name to the list
    return healthBoardsList


def getNewest():
    """
    Reads the covid data and returns the total cases for all health boards
    :return: (list): A list of each health boards covid numbers
    """
    newDataList = []                                        # The list to store the case values
    for col in range(2, 17):                                # Loop each health board area & Scottish total
        cell = sheet.cell(row=sheet.max_row, column=col)    # Store the cell for the current health board on the row
        newDataList.append(cell.value)                      # Get the data for that cell  (stores case numbers)
    return newDataList                                      # Returns all the case values in the list


def getScotlandTotal():
    """
    Reads the covid data to find the total cases for Scotland
    :return: (str): The total cases in Scotland as a String
    """
    # Reads the last column of data, on the last row of data, and returns that cells value
    return sheet.cell(row=sheet.max_row, column=16).value


def getHealthBoardTotal(healthBoard):
    """
    Takes a health boards name, returning its total cases
    :param healthBoard: (str): The name of the health board to check for
    :return: (str): The number of cases from the excel sheet
    """
    columnNum = getHealthBoardColumnNum(healthBoard)    # Gets the column number, for the given health board
    newData = getNewest()                 # Stores all of the newest available total cases in a list
    return newData[columnNum]                   # From the list of data, select the element from the given column num


def getHealthBoardColumnNum(healthBoard):
    """
    Takes the health board name, returning its column number
    :param healthBoard: (str): The name of the health board
    :return: (int): The column number from the excel sheet
    """
    for col in range(2, 17):            # Loops all of the columns of health boards
        if sheet.cell(row=3, column=col).value == healthBoard:  # Check if the current health board matches the request
            # Remove 2 from the column number, as the column in excel, starts at 2
            # so removing 2 the first element back to 1, so it becomes the 'first' column again
            return col - 2              # Returns the column number


def getHealthBoardPeriod(timePeriod, healthBoard='all'):
    """
    Reads the covid data for the health boards covid cases, over the given period
    :param healthBoard: (str): The name of the health board to check data for. Use 'all' to get all health boards cases
    :param timePeriod: (int): The amount of days of data to check back for
    :return: (list): A list of the cases from either all health boards or the specifically requested one
    """
    # 39 was selected due to formatting in the Gov file
    # When cases were less than 5, as '*' was displayed, for disclosure reasons
    # So, to avoid a ValueError, 39 was the first row of data to not have an *
    # TODO: Find a way to accommodate the asterix value
    max = sheet.max_row - 39
    try:
        length = int(timePeriod)
    except:
        print('ERROR: You must give a number between 1 and ' + str(max) + '\nEnding Program.')
        sys.exit()
    if int(length) < 1 or int(length) > max:
        print('ERROR: Given value is not valid for data available. Please enter a value between 1 and ' + str(max))

    if healthBoard == 'all':
        print('Getting all health boards cases over ' + str(length) + ' days')  # TODO Remove this after list handling
        healthBoardList = []
        for col in range(2, 17):
            newCell = sheet.cell(row=sheet.max_row, column=col).value
            olderCell = sheet.cell(row=sheet.max_row - length, column=col).value
            data = newCell - olderCell
            healthBoardList.append(data)
        return healthBoardList
    else:
        healthBoard = getHealthBoardFullName(healthBoard)
        print('Getting the last ' + str(length) + ' days of cases for ' + healthBoard)  # TODO Remove after CMD handling
        healthBoardColNum = getHealthBoardColumnNum(healthBoard)
        newCell = sheet.cell(row=sheet.max_row, column=healthBoardColNum).value
        olderCell = sheet.cell(row=sheet.max_row - length, column=healthBoardColNum).value
        return newCell - olderCell


def outputHandler(locations, values):
    """
    Ouputs the health board and case numbers in a tabulated list
    :param locations: (list or str): A list of all health boards or a single health board
    :param values: (str): The number of cases for the health board(s)
    """
    # Checks if there is a list of health boards or just a single health board
    if type(locations) == list and type(values) == list:
        print('Received a list, so it should be ALL health boards')     # TODO: Remove these prints
        maxLenElement = max(locations, key=len)                         # Store the longest element from the list
        for x in range(len(locations)):
            # Take the current items length, remove it from the length of the longest element, add 2 for the clarity
            spacing = ' ' * (len(maxLenElement) - len(locations[x]) + 2)
            print(locations[x] + spacing + ' | ' + str(values[x]))
    elif type(locations) == str and type(values) == int:
        print('Received a str, so it will just be a single health board')
        print(locations + '\t|\t' + str(values))                        # Tab twice to add some space between values


def getHealthBoardFullName(location):
    """
    Takes user inputted health board name, converting it into the name as it appears in the covid data excel sheet
    User input: Grampian, returns NHS Grampian
    :param location: (str): The health board name that
    :return: (str): The official name of the health board
    """
    locationsList = getHealthBoardList()
    healthBoardNameFull = ''                        # Stores the health board name
    for x in range(len(locationsList)):             # Loops the entire list of health boards
        if location.lower() in locationsList[x].lower():    # If the given finds a match in the list of locations
            healthBoardNameFull = locationsList[x]  # Store that full health board name in a variable
    if healthBoardNameFull == '':                   # If there is no match, print error, print valid names and end
        print('Error - Invalid name given, please enter a name that matches one of the following:')
        print(getHealthBoardList())                 # Prints the health board names
        print('Ending program')
        sys.exit()                                  # Uses the sys package to end the program, as no valid name given
    elif healthBoardNameFull in locationsList:      # If there is a matching name in the location list, return as String
        return healthBoardNameFull

# TODO; Set up argparse


newestFile = formatFileName()  # Stores the expected file name from the website
today = getFormattedDate()  # Gets the date in a URL format to add to the source file URL
# The URL of where the most recent file will be
fileURL = "http://www.gov.scot/binaries/content/documents/govscot/publications/statistics/2020/04/coronavirus" \
              "-covid-19-trends-in-daily-data/documents/covid-19-data-by-nhs-board/covid-19-data-by-nhs-board/govscot" \
              "%3Adocument/COVID-19%2Bdaily%2Bdata%2B-%2Bby%2BNHS%2BBoard%2B-%2B" + today + ".xlsx?forceDownload=true "

# A file with the most recent data does not already exist
if newestFile not in os.listdir(os.getcwd() + '\\ExcelFiles\\'):
    print('Local covid data is out of date - Downloading recent data.')
    downloadData(fileURL, newestFile)    # Downloads newest file is available or uses older file

# File management - Clear out any other older files
count = 0
for theFile in os.listdir(os.getcwd() + '\\ExcelFiles\\'):
    if theFile != newestFile:
        count += 1
        # print('need to bin this file: ' + str(os.getcwd() + '\\ExcelFiles\\' + theFile))
        send2trash(os.getcwd() + '\\ExcelFiles\\' + theFile)

# Loads the most recent excel file, data_only used to ignore any formulas, as we only need the actual values
excel = load_workbook('ExcelFiles//' + newestFile, data_only=True)

# Get the correct sheet from the Excel file
for theSheet in range(len(excel.sheetnames)):
    if excel.sheetnames[theSheet] == 'Table 1 - Cumulative cases':
        excel.active = theSheet             # The active sheet is used to access the sheets data
sheet = excel.active

lastRowNum = sheet.max_row
newestDate = sheet.cell(row=lastRowNum, column=1)  # Get the data of the most available row of data

print(' ---- Scottish Covid Case Checker ---- ')
# TODO: Is this needed? urllib.request.urlcleanup() - https://docs.python.org/3/library/urllib.request.html
