#! python3
# ScottishCovidCases.py - A program to check the new coronavirus cases in Scotland
# Data is taken from - https://www.gov.scot/publications/coronavirus-covid-19-trends-in-daily-data/

# Imports
from bs4 import BeautifulSoup as bs
import requests
from datetime import date
import urllib.request
from openpyxl import *
import os
import sys
import shutil
from send2trash import send2trash
import argparse
import re
import platform


# Functions
def getFormattedDate(formatted=True):
    """
    Returns the date in a specified format, depending on parameters
    :param formatted: If True, returns the data formatted for a URL, otherwise its just DDMonthYYYY
    :return: (str): A String representing the formatted data
    """
    # Get date in DDMMYYYY format
    today = date.today()  # Create an object to get the date details
    dayNum = today.strftime("%d")  # Returns today's date in a int
    month = today.strftime("%B")  # Returns today's month in word format
    year = today.strftime("%Y")  # Returns year in 4 digits
    # Check if formatting is required or not, and return the value as a String
    if formatted:
        # Returns data, formatted to be used as part of a URL
        dateStr = dayNum + "%2B" + month + "%2B" + year
    else:
        # Returns today's date in format - DDMonthYear, e.g. 25October2020
        dateStr = dayNum + month + year
    return dateStr


def downloadData(url, file):
    """
    Downloads the latest excel file of Scottish Covid cases from the given URL
    :param url: (str): The URL of the Scottish Gov Covid Cases website
    :param file: (str; The name of the file to check for
    """
    # Check if a file with that name already exists, if not, download a fresh copy
    # Uses the os module to check the name of the file to download, against all the files in the ExcelFiles directory
    if platform.system() == "Windows":
        dir = r"\\ExcelFiles\\"
    else:
        dir = '/ExcelFiles/'
    print(os.getcwd() + dir)
    if file in os.listdir(os.getcwd() + dir):
        print('A file with today\'s date already exists, using that.')
    else:
        try:
            # Attempts to use the urllib module to download the given file from the Scot Gov website
            urllib.request.urlretrieve(url, file)  # Downloads to the same directory as the python file
        # If there is an issue with the given URL, display an error, the given URL and end the program
        except urllib.error.URLError:
            print("Error! The given URL is incorrect, please check the URL")
            print("URL Given: " + url)
            print('Ending program as no data available to analyse')
            sys.exit()  # Ends the program as the URL failed, so no data available
        print("Local data matches most recently available data")
        # Move the newly downloaded file into the correct directory
        shutil.move(newestFile, 'ExcelFiles')


def getURLs():
    """
    Scans a web page for all URLs
    :return: (list); A list of all of the URLS from the Scottish Gov Covid page
    """
    # The website that hosts the files
    url = 'https://www.gov.scot/publications/coronavirus-covid-19-trends-in-daily-data/'
    # Create a soup object, that reads the HTML content and create a list to later store the results
    soup = bs(requests.get(url).text, 'html.parser')
    links = []

    # Loop through all of the content in the soup object, adding each hyperlink to the list and return the list
    # This allows the hyperlinks to be searched for the correct file later, by 'etFileNameFromLinks()
    for a in soup.find_all('a'):
        links.append(a['href'])
    return links


def formatFileName():
    """
    Gets the name of the covid data sheet and formats it to rename the exhcel sheet when its downloaded
    :return: (str): A String to name the downloaded excel sheet of covid data
    """
    links = getURLs()  # Get a list of URLs to search through for the Excel file
    fileName = ''  # The intended filename for the download object

    try:
        # Loop through all of the URLs in the list, for the specific file name and store in the fileName variable
        for link in links:
            if '/binaries/content/documents/govscot/publications/statistics/2020/04/coronavirus-covid-19-trends-in' \
               '-daily-data/documents/covid-19-data-by-nhs-board/' in link:
                if '?forceDownload' in link:
                    fileName = link

        stringPos = fileName.index('COVID-19%2B')  # Finds the start of the file name and stores its character position
        # Removes the first part of the URL to get a cleaner file name, using the above index as the cut off point
        fileName = fileName[stringPos:]
        # Removes all the characters from the String that are not needed (cleaning it up)
        fileName = fileName.replace('%2B', '')
        fileName = fileName.replace('?forceDownload=true', '')
        fileName = fileName.replace('dailydata-byNHSBoard', '')
        # The file name in a readable format
        return fileName
    except:
        # Print warning message to user
        print('Error! Found no file to download. Please check the URL has a valid file to download.\nIt is possible '
              'the file name has changed.')
        sys.exit()


def getHealthBoardList():
    """
    Reads the covid data and returns all of the health boards
    :return: (list): A list of each Health Board name as a Strings
    """
    healthBoardsList = []  # Used to store all the health board names
    # Loops from the first available health board name (col 2) to the last health board name (col 16)
    for row in sheet.iter_rows(min_row=3, min_col=2, max_row=3, max_col=16):
        for cell in row:  # Get the cell of the iteration
            healthBoardsList.append(cell.value)  # Add the health board name to the list
    return healthBoardsList


def getNewest():
    """
    Reads the covid data and returns the total cases for all health boards
    :return: (list): A list of each health boards covid numbers
    """
    newDataList = []  # The list to store the case values
    for col in range(2, 17):  # Loop each health board area & Scottish total
        cell = sheet.cell(row=lastRowNum, column=col)  # Store the cell for the current health board on the row
        newDataList.append(cell.value)  # Get the data for that cell  (stores case numbers)
    return newDataList  # Returns all the case values in the list


def getScotlandTotal():
    """
    Reads the covid data to find the total cases for Scotland
    :return: (str): The total cases in Scotland as a String
    """
    # Reads the last column of data, on the last row of data, and returns that cells value
    data = sheet.cell(row=lastRowNum, column=16).value
    if type(data) is not int:
        data = int(re.sub("[^0-9]", "", data))
    return int(data)


def getHealthBoardTotal(healthBoard):
    """
    Takes a health boards name, returning its total cases
    :param healthBoard: (str): The name of the health board to check for
    :return: (str): The number of cases from the excel sheet
    """
    columnNum = getHealthBoardColumnNum(healthBoard)  # Gets the column number, for the given health board
    newData = getNewest()  # Stores all of the newest available total cases in a list
    data = newData[columnNum - 2]
    if type(data) is not int:
        data = int(re.sub("[^0-9]", "", data))
    return data  # From the list of data, select the element from the given column num


def getHealthBoardColumnNum(healthBoard):
    """
    Takes the health board name, returning its column number
    :param healthBoard: (str): The name of the health board
    :return: (int): The column number from the excel sheet
    """
    for col in range(2, 17):  # Loops all of the columns of health boards
        if sheet.cell(row=3, column=col).value == healthBoard:  # Check if the current health board matches the request
            # Remove 2 from the column number, as the column in excel, starts at 2
            # so removing 2 the first element back to 1, so it becomes the 'first' column again
            return col  # Returns the column number


def getHealthBoardPeriod(timePeriod, healthBoard='all'):
    """
    Reads the covid data for the health boards covid cases, over the given period
    :param healthBoard: (str): The name of the health board to check data for. Use 'all' to get all health boards cases
    :param timePeriod: (int): The amount of days of data to check back for
    :return: (list): A list of the cases from either all health boards or the specifically requested one
    """
    # 39 was selected due to formatting in the Gov file
    # When cases were less than 5, as '*' was displayed, for disclosure reasons
    # So, to avoid a ValueError, 39 was the first row of data to not have an *
    maxRow = lastRowNum - 39
    try:
        length = int(timePeriod)
    except ValueError:
        print('ERROR: You must give a number between 1 and ' + str(maxRow) + '\nEnding Program.')
        sys.exit()
    if int(length) < 1 or int(length) > maxRow:
        print('ERROR: Given days value is not valid for the data available. Please enter a days value between 1 and ' + str(maxRow))
        sys.exit()

    if healthBoard == 'all':
        print('Getting all health boards cases over ' + str(length) + ' days')
        healthBoardList = []
        for col in range(2, 17):
            newCell = sheet.cell(row=lastRowNum, column=col).value
            olderCell = sheet.cell(row=lastRowNum - length, column=col).value
            if type(newCell) is not int:
                newCell = int(re.sub("[^0-9]", "", newCell))
            if type(olderCell) is not int:
                olderCell = int(re.sub("[^0-9]", "", olderCell))
            data = int(newCell) - int(olderCell)
            healthBoardList.append(data)
        return healthBoardList
    else:
        print('Getting the last ' + str(length) + ' days of cases for ' + healthBoard)
        healthBoardColNum = getHealthBoardColumnNum(healthBoard)
        newCell = sheet.cell(row=lastRowNum, column=healthBoardColNum).value
        olderCell = sheet.cell(row=lastRowNum - length, column=healthBoardColNum).value
        return newCell - olderCell


def outputData(locations, values):
    """
    Ouputs the health board and case numbers in a tabulated list
    :param locations: (list or str): A list of all health boards or a single health board
    :param values: (str): The number of cases for the health board(s)
    """
    # Checks if there is a list of health boards or just a single health board
    if type(locations) == list and type(values) == list:
        maxLenElement = max(locations, key=len)  # Store the longest element from the list
        for x in range(len(locations)):
            # Take the current items length, remove it from the length of the longest element, add 2 for the clarity
            spacing = ' ' * (len(maxLenElement) - len(locations[x]) + 2)
            print(locations[x] + spacing + ' | ' + str(values[x]))
    elif type(locations) == str and type(values) == int:
        print(locations + '\t|\t' + str(values))  # Tab twice to add some space between values


def getHealthBoardFullName(location):
    """
    Takes user inputted health board name, converting it into the name as it appears in the covid data excel sheet
    User input: Grampian, returns NHS Grampian
    :param location: (str): The health board name that
    :return: (str): The official name of the health board
    """
    valid = True
    healthBoardNameFull = ''  # Stores the health board name
    if valid:
        locationsList = getHealthBoardList()
        for x in range(len(locationsList)):  # Loops the entire list of health boards
            splitLocation = locationsList[x].split()
            for y in range(len(splitLocation)):
                if location.lower() == splitLocation[y].lower():
                    healthBoardNameFull = locationsList[x]
                    break
            if healthBoardNameFull != '':  # Escapes the FOR loop, as there is now a health board name
                break
            # if location.lower() in locationsList[x].lower():  # If the given finds a match in the list of locations
            #     healthBoardNameFull = locationsList[x]  # Store that full health board name in a variable
            #     break
    if healthBoardNameFull == '':  # If there is no match, print error, print valid names and end
        print('Error - Invalid name given, please enter a name that matches one of the following:')
        print(healthBoards)  # Prints the health board names
        print('Ending program')
        sys.exit()  # Uses the sys package to end the program, as no valid name given
    elif healthBoardNameFull in locationsList:  # If there is a matching name in the location list, return as String
        return healthBoardNameFull


def handleInput(theArea):
    """
    Takes the command line arguments health board name and returns the expected Health Board name
    :param theArea: The command line arguments passed
    :return: The Health board name, that matches the same name in hte Excel data
    """
    areaName = ''
    for namePart in theArea:
        if len(areaName) < 1:
            if namePart != "&" and namePart != "NHS":
                for x in getHealthBoardList():
                    if namePart in x and namePart != "&" and namePart != "NHS":
                        areaName = getHealthBoardFullName(str(namePart))
                        return areaName
        else:
            return areaName


# argparse variables
parser = argparse.ArgumentParser(prog=os.path.basename(__file__), usage='%(prog)s [option]',
                                 description='---- Scottish Covid Case Checker ---- \nAnalyses Scottish Covid-19 case '
                                             'numbers and returns specific case numbers',
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
group = parser.add_mutually_exclusive_group()
group.add_argument('-n', '--new', required=False, action='store_true', help="Returns todays newest case numbers for "
                                                                            "each health board")
group.add_argument('-s', '--scotland', required=False, action='store_true',
                   help="Returns the Scottish total amount of cases")
group.add_argument('-a', '--area', required=False, nargs='*', help="Takes a health board name, returns that health "
                                                                   "boards total cases",
                   metavar='HEALTHBOARD')
group.add_argument('-c', '--cases', required=False, nargs='*', help="Takes a number of days & a health board or "
                                                                    "\'all\', returns the case numbers over that "
                                                                    "period",
                   metavar=('DAYS', 'HEALTHBOARD'))
group.add_argument('-t', '--total', required=False, action='store_true',
                   help="Returns all health boards total case numbers")
group.add_argument('-hb', '--healthboards', required=False, action='store_true',
                   help="Returns all health boards available")
args = parser.parse_args()

newestFile = formatFileName()  # Stores the expected file name from the website
today = getFormattedDate()  # Gets the date in a URL format to add to the source file URL
fileURL = "http://www.gov.scot/binaries/content/documents/govscot/publications/statistics/2020/04/coronavirus" \
          "-covid-19-trends-in-daily-data/documents/covid-19-data-by-nhs-board/covid-19-data-by-nhs-board/govscot" \
          "%3Adocument/COVID-19%2Bdaily%2Bdata%2B-%2Bby%2BNHS%2BBoard%2B-%2B" + today + ".xlsx?forceDownload=true "

healthBoards = ['Ayrshire Arran', 'Borders', 'Dumfries Galloway', 'Fife', 'Forth Valley', 'Grampian', 'Greater Glasgow Clyde', 'Highland', 'Lanarkshire', 'Lothian', 'Orkney', 'Shetland', 'Tayside', 'Western Isles', 'Scotland']

if platform.system() == "Windows":
    excelDir = "\\ExcelFiles\\"
else:
    excelDir = '/ExcelFiles/'

# Checks if their is a suitable directory to store the Excel files, if not, makes one
if 'ExcelFiles' not in os.listdir(os.getcwd()):
    os.mkdir(os.getcwd() + excelDir)
# A file with the most recent data does not already exist
if newestFile not in os.listdir(os.getcwd() + excelDir):
    print('Local covid data is out of date - Downloading recent data.')
    downloadData(fileURL, newestFile)  # Downloads newest file is available or uses older file
    urllib.request.urlcleanup()

# File management - Clear out any older Excel files
count = 0
for theFile in os.listdir(os.getcwd() + excelDir):
    if theFile != newestFile:
        count += 1
        send2trash(os.getcwd() + excelDir + theFile)

# Loads the most recent excel file, data_only used to ignore any formulas, as we only need the actual values
if platform.system() == "Windows":
    excel = load_workbook('ExcelFiles//' + newestFile, data_only=True)
else:
    excel = load_workbook('ExcelFiles\\' + newestFile, data_only=True)

# Get the correct sheet from the Excel file
for theSheet in range(len(excel.sheetnames)):
    if excel.sheetnames[theSheet] == 'Table 1 - Cumulative cases':
        excel.active = theSheet  # The active sheet is used to access the sheets data
sheet = excel.active

# Sometimes the sheet comes back with the last row, as the one after the last line of actual data
# So, instead of using sheet.max_row, loop from the last row, until the first row of actual data, store that row number
lastEmptyRow = sheet.max_row
while sheet.cell(row=lastEmptyRow, column=1).value is None:
    lastEmptyRow -= 1
lastRowNum = lastEmptyRow

# A message to display during certain CLI arguments
intro = '\n---- Scottish Covid Case Checker ---- \nAnalyses Scottish Covid-19 cases and returns specific case numbers\n'

# CLI Input handler
# Uses argparse arguments to run a specific function
if args.new is True:
    # Returns the newest case numbers, 1 only shows the cases added since yesterdays data
    print(intro)
    outputData(getHealthBoardList(), getHealthBoardPeriod(1, 'all'))
elif args.scotland is True:
    print(intro)
    print('Scotlands total cases')
    outputData(getHealthBoardFullName('Scotland'), getScotlandTotal())
elif args.area is not None:
    print(intro)
    area = handleInput(args.area)
    if area in getHealthBoardList():
        print(area + 's total cases')
        outputData(area, getHealthBoardTotal(area))
    else:
        print('Error - Invalid name given, please enter a name that matches one of the following:')
        print(healthBoards)  # Prints the health board names
        print('Ending program')
    sys.exit()
elif args.cases is not None:
    # Takes the given area, converts to Excel formatted health board name and checks the the cases for the given days
    print(intro)
    try:
        if args.cases[1] != 'all':
            onlyArea = args.cases[1:]
            area = handleInput(onlyArea)
            if area in getHealthBoardList():
                outputData(area, getHealthBoardPeriod(args.cases[0], area))
            else:
                print('Error - Invalid name given, please enter a name that matches one of the following:')
                print(healthBoards)  # Prints the health board names
                print('Ending program')
        else:
            area = 'all'
            outputData(getHealthBoardList(), getHealthBoardPeriod(args.cases[0], 'all'))
    except IndexError as e:
        print('Error: Missing the health board name argument')
        print('-c requires a number for days (e.g. 1, 7, 2) and a Health Board name or all')
elif args.total is True:
    # Prints all health boards and all of the total case numbers
    print(intro)
    print('Every health boards total case numbers')
    outputData(getHealthBoardList(), getNewest())
elif args.healthboards is True:
    print(intro)
    print('The following Health Boards can be used as arguments:')
    # Outputs the health board list in an actual visual list
    for hb in healthBoards:
        print('* ', hb)
    sys.exit()
else:
    # Invalid argument selected, showing the user -h
    parser.print_help()
